def excel_to_pdf(name):
    from openpyxl import load_workbook
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter

    def convert_excel_to_pdf(input_excel, output_pdf):
        # Avaa Excel-tiedosto
        workbook = load_workbook(input_excel)
        sheet = workbook.active
        # Luo PDF-tiedosto
        pdf = canvas.Canvas(output_pdf, pagesize=letter)
        # Määritä sarakkeiden ja rivien määrä
        rows = sheet.max_row
        cols = sheet.max_column
        # Määritä sarakkeiden leveys ja rivien korkeus (voit säätää tarpeen mukaan)
        col_widths = [50 for _ in range(cols)]
        row_height = 15
        # Määritä PDF-sivun marginaalit
        margin = 20
        width, height = letter
        usable_width = width - 2 * margin
        # Laske sarakkeiden leveys suhteessa käytettävissä olevaan leveyteen
        col_widths = [width * (col_width / sum(col_widths)) for col_width in col_widths]
        # Tulosta Excel-tiedoston sisältö PDF:ään
        for row in range(1, rows + 1):
            y = height - margin - row * row_height
            for col in range(1, cols + 1):
                x = margin + sum(col_widths[:col - 1])
                cell_value = str(sheet.cell(row=row, column=col).value)
                pdf.drawString(x, y, cell_value)
            # Älä kutsu showPage() joka rivin jälkeen
        pdf.showPage()
        # Tallenna PDF-tiedosto
        pdf.save()
    # if __name__ == "__main__":
    input_excel_path = f'{name}.xlsx'
    output_pdf_path = f'{name}.pdf'
    convert_excel_to_pdf(input_excel_path, output_pdf_path)


def python_list_to_excel(python_list,name):
    ### TO EXCEL Start ###
    import pandas as pd
    from openpyxl import Workbook
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
    from io import BytesIO
    df = pd.DataFrame(python_list)
    # Tallenna DataFrame Excel-tiedostoon
    excel_file = f'{name}.xlsx'
    df.to_excel(excel_file, index=False)
    ### TO EXCEL END ###